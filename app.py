import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import io
import math
import base64
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# PAY STRUCTURE
# ──────────────────────────────────────────────────────────────────────────────
PAY_STRUCTURE = {
    "A6": {"score": 0.6, "salary": 2625.09, "mobility": 780},
    "A7": {"score": 0.7, "salary": 2756.35, "mobility": 780},
    "A8": {"score": 0.8, "salary": 2894.16, "mobility": 780},
    "A9": {"score": 0.9, "salary": 3038.87, "mobility": 780},
    "B0": {"score": 1.0, "salary": 3190.81, "mobility": 930},
    "B1": {"score": 1.1, "salary": 3350.35, "mobility": 930},
    "B2": {"score": 1.2, "salary": 3517.87, "mobility": 930},
    "B3": {"score": 1.3, "salary": 3693.77, "mobility": 930},
    "B4": {"score": 1.4, "salary": 3878.45, "mobility": 930},
    "B5": {"score": 1.5, "salary": 4072.38, "mobility": 930},
    "B6": {"score": 1.6, "salary": 4276.00, "mobility": 930},
    "B7": {"score": 1.7, "salary": 4489.80, "mobility": 930},
    "B8": {"score": 1.8, "salary": 4714.29, "mobility": 930},
    "B9": {"score": 1.9, "salary": 4950.00, "mobility": 930},
    "C0": {"score": 2.0, "salary": 5197.50, "mobility": 1080},
    "C1": {"score": 2.1, "salary": 5457.38, "mobility": 1080},
    "C2": {"score": 2.2, "salary": 5730.24, "mobility": 1080},
    "C3": {"score": 2.3, "salary": 6016.76, "mobility": 1080},
    "C4": {"score": 2.4, "salary": 6317.59, "mobility": 1080},
    "C5": {"score": 2.5, "salary": 6633.47, "mobility": 1080},
    "C6": {"score": 2.6, "salary": 6965.15, "mobility": 1080},
    "C7": {"score": 2.7, "salary": 7243.75, "mobility": 1080},
    "C8": {"score": 2.8, "salary": 7533.50, "mobility": 1080},
    "C9": {"score": 2.9, "salary": 7834.84, "mobility": 1080},
    "D0": {"score": 3.0, "salary": 8148.24, "mobility": 1230},
    "D1": {"score": 3.1, "salary": 8474.17, "mobility": 1230},
    "D2": {"score": 3.2, "salary": 8813.13, "mobility": 1230},
    "D3": {"score": 3.3, "salary": 9165.66, "mobility": 1230},
    "D4": {"score": 3.4, "salary": 9532.28, "mobility": 1230},
    "D5": {"score": 3.5, "salary": 9913.58, "mobility": 1230},
    "D6": {"score": 3.6, "salary": 10310.12, "mobility": 1230},
    "D7": {"score": 3.7, "salary": 10722.52, "mobility": 1230},
    "D8": {"score": 3.8, "salary": 11151.42, "mobility": 1230},
    "D9": {"score": 3.9, "salary": 11597.48, "mobility": 1230},
    "E0": {"score": 4.0, "salary": 12061.38, "mobility": 1380},
    "E1": {"score": 4.1, "salary": 12543.84, "mobility": 1380},
    "E2": {"score": 4.2, "salary": 13045.59, "mobility": 1380},
    "E3": {"score": 4.3, "salary": 13567.41, "mobility": 1380},
    "E4": {"score": 4.4, "salary": 14110.11, "mobility": 1380},
    "E5": {"score": 4.5, "salary": 14674.51, "mobility": 1380},
}

SCORE_LABELS = {
    0: "0 — Not present",
    1: "1 — Basic",
    2: "2 — Developing",
    3: "3 — Proficient",
    4: "4 — Advanced",
    5: "5 — Expert",
}

PAGES = [
    ("Job Information",       "job_info"),
    ("Technical Competency",  "technical"),
    ("Behavioural Competency","behavioural"),
    ("Effort",                "effort"),
    ("Professional Capital",  "professional"),
    ("Working Conditions",    "working"),
    ("Responsibility",        "responsibility"),
    ("Results & Pay Proposal","results"),
    ("Reward Strategy",       "info"),
]

# ──────────────────────────────────────────────────────────────────────────────
# SESSION STATE — initialise all keys once so widgets never reset on navigate
# ──────────────────────────────────────────────────────────────────────────────
SCORE_KEYS = [
    "tc_legal", "tc_data", "tc_strategy", "tc_leadership", "tc_transformational",
    "bc_ic_cp", "bc_ic_cs", "bc_ic_team", "bc_ic_org", "bc_freq", "bc_cons", "bc_conf",
    "ef_conc", "ef_prob", "ef_info", "ef_multi", "ef_switch",
    "ef_own", "ef_oth", "ef_conf", "ef_press",
    "pc_cred", "pc_rel", "pc_org",
    "wc_sched", "wc_travel", "wc_social",
    "resp_scope", "resp_auto", "resp_rev", "resp_dec",
]

# Only Technical Competency and Professional Capital allow a score of 0,
# with one exception: Legal has a floor of 1 because Stratarius is fundamentally
# a legal advisory firm — at least one technical sub-dimension must therefore
# score ≥ 1. This guarantees a minimum raw score of 0.650 (→ A6 in the pay
# structure).
ZERO_OK_KEYS = {
    "tc_data", "tc_strategy", "tc_leadership", "tc_transformational",
    "pc_cred", "pc_rel", "pc_org",
}

def min_score(key):
    return 0 if key in ZERO_OK_KEYS else 1

DEFAULTS = {
    # Job info
    "job_name": "", "evaluator": "",
    # Each score starts at its lowest allowed value (0 for TC/PC, 1 elsewhere)
    **{k: min_score(k) for k in SCORE_KEYS},
    # Per-sub-dimension reasoning (optional, free text)
    **{f"comment_{k}": "" for k in SCORE_KEYS},
    # Overall reasoning on Results page
    "overall_comments": "",
    # Navigation
    "_page": "job_info",
}

def init_state():
    for k, v in DEFAULTS.items():
        if k not in st.session_state:
            st.session_state[k] = v

# ──────────────────────────────────────────────────────────────────────────────
# GLOBAL STYLES
# ──────────────────────────────────────────────────────────────────────────────
STYLES = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
}

/* ── Force light app background (overrides any dark base theme) ── */
.stApp, [data-testid="stAppViewContainer"], .main, [data-testid="stMain"] {
    background-color: #FFFFFF !important;
}
[data-testid="stHeader"] { background: transparent !important; }

/* ── Sidebar permanently pinned open ── */
[data-testid="stSidebar"] {
    min-width: 244px !important;
    max-width: 244px !important;
    transform: none !important;
    visibility: visible !important;
    margin-left: 0 !important;
}
[data-testid="stSidebar"][aria-expanded="false"] {
    margin-left: 0 !important;
    transform: none !important;
}
/* Hide every flavour of sidebar collapse/expand button so it can't be toggled away */
[data-testid="stSidebarCollapseButton"],
[data-testid="stSidebarCollapsedControl"],
[data-testid="collapsedControl"],
button[kind="header"],
button[kind="headerNoPadding"] {
    display: none !important;
}

/* ── Sidebar — light theme ── */
[data-testid="stSidebar"] {
    background: #FFFFFF !important;
    border-right: 1px solid #D8EBE7 !important;
}
[data-testid="stSidebar"] * { color: #164A41 !important; }
[data-testid="stSidebarContent"] { padding: 0 !important; }

.sidebar-logo {
    background: #FFFFFF;
    padding: 22px 20px 18px 20px;
    border-bottom: 1px solid #D8EBE7;
    margin-bottom: 6px;
    text-align: left;
}
.sidebar-logo img { height: 36px; display: block; }
.sidebar-brand {
    color: #4A7A70 !important;
    font-size: 11px;
    font-weight: 600;
    letter-spacing: 0.1em;
    text-transform: uppercase;
    margin-top: 12px;
    text-align: left;
}

/* Nav buttons */
[data-testid="stSidebar"] .stButton > button {
    background: transparent !important;
    border: none !important;
    border-radius: 7px !important;
    color: #4A7A70 !important;
    font-size: 13.5px !important;
    font-weight: 500 !important;
    text-align: left !important;
    justify-content: flex-start !important;
    padding: 9px 14px !important;
    width: 100% !important;
    transition: background 0.15s, color 0.15s !important;
    box-shadow: none !important;
}
[data-testid="stSidebar"] .stButton > button > div,
[data-testid="stSidebar"] .stButton > button p {
    text-align: left !important;
    width: 100% !important;
}
.score-preview, .score-preview-label, .score-big, .score-level, .score-salary { text-align: left !important; }
[data-testid="stSidebar"] .stButton > button:hover {
    background: #EFF7F4 !important;
    color: #164A41 !important;
}
[data-testid="stSidebar"] .stButton > button[kind="primary"] {
    background: #E8F4F1 !important;
    border-left: 3px solid #164A41 !important;
    color: #164A41 !important;
    font-weight: 600 !important;
}

.score-preview {
    margin: 14px 10px 6px 10px;
    background: #F5FAF8;
    border-radius: 10px;
    padding: 14px 16px;
    border: 1px solid #D8EBE7;
}
.score-preview-label {
    font-size: 10px !important;
    font-weight: 600;
    letter-spacing: 0.1em;
    text-transform: uppercase;
    color: #7FA89F !important;
    margin-bottom: 8px;
}
.score-big { font-size: 30px !important; font-weight: 700 !important; color: #164A41 !important; line-height: 1.1; }
.score-level { font-size: 14px !important; font-weight: 600 !important; color: #E07B39 !important; margin-top: 2px; }
.score-salary { font-size: 11px !important; color: #7FA89F !important; margin-top: 3px; }
.score-bar-bg { background: #D8EBE7; border-radius: 4px; height: 4px; margin-top: 10px; overflow: hidden; }
.score-bar-fill { height: 4px; border-radius: 4px; background: linear-gradient(90deg,#164A41,#E07B39); }

/* ── Main content ── */
.main .block-container { padding: 2rem 2.5rem 3rem 2.5rem; max-width: 960px; }

/* ── Page header ── */
.page-header { margin-bottom: 26px; padding-bottom: 18px; border-bottom: 2px solid #D8EBE7; }
.page-header-badge {
    display: inline-flex; align-items: center;
    background: #E8F4F1; color: #164A41;
    font-size: 11px; font-weight: 700; letter-spacing: 0.1em; text-transform: uppercase;
    padding: 3px 10px; border-radius: 20px; margin-bottom: 10px;
}
.page-title {
    font-size: 27px !important; font-weight: 700 !important;
    color: #164A41 !important; margin: 0 0 5px 0 !important; letter-spacing: -0.4px;
}
.page-subtitle { font-size: 15px; color: #64748B; margin: 0; }

/* ── Cards ── */
.card {
    background: #FFFFFF; border: 1px solid #D8EBE7; border-radius: 12px;
    padding: 20px 24px; margin-bottom: 16px;
    box-shadow: 0 1px 3px rgba(22,74,65,0.06);
}
.card-title {
    font-size: 11px; font-weight: 700; letter-spacing: 0.08em; text-transform: uppercase;
    color: #94A3B8; margin-bottom: 14px; padding-bottom: 10px; border-bottom: 1px solid #EDF2F7;
}

/* ── Info panel & radio cards share the same outer width and box model ── */
.info-panel, div[data-testid="stRadio"] > div > label {
    width: 100% !important;
    box-sizing: border-box !important;
}
.info-panel {
    background: #EFF7F4; border: 1px solid #B8D8D2; border-left: 4px solid #164A41;
    border-radius: 0 8px 8px 0; padding: 14px 18px; margin-bottom: 22px;
    font-size: 15px; color: #164A41; line-height: 1.6;
}
/* The Streamlit radio wrapper sometimes inherits a width:auto from baseweb; force it
   to fill the column so each option card lines up with the info panel above. */
div[data-testid="stRadio"], div[data-testid="stRadio"] > div { width: 100% !important; }

/* ── Metric cards ── */
.metric-card {
    background: #FFFFFF; border: 1px solid #D8EBE7; border-radius: 10px;
    padding: 16px 18px; text-align: center;
    box-shadow: 0 1px 3px rgba(22,74,65,0.06); height: 100%;
}
.metric-label { font-size: 11px; font-weight: 600; letter-spacing: 0.07em; text-transform: uppercase; color: #94A3B8; margin-bottom: 6px; }
.metric-value { font-size: 26px; font-weight: 700; color: #164A41; line-height: 1.1; }
.metric-highlight { border: 2px solid #E07B39 !important; }
.metric-highlight .metric-value { color: #E07B39; }

/* ── Pay level badge ── */
.pay-level-badge {
    display: inline-block;
    background: linear-gradient(135deg,#164A41 0%,#1D5C4E 100%);
    color: #FFFFFF; font-size: 22px; font-weight: 700;
    padding: 6px 18px; border-radius: 8px; letter-spacing: 2px;
}

/* ── Dimension score row ── */
.dim-row { display: flex; align-items: center; padding: 9px 0; border-bottom: 1px solid #F1F5F9; }
.dim-row:last-child { border-bottom: none; }
.dim-name { flex: 1; font-size: 13px; color: #334155; font-weight: 500; }
.dim-weight { font-size: 11px; color: #94A3B8; width: 48px; text-align: right; margin-right: 12px; }
.dim-score-bar { flex: 0 0 110px; }
.dim-score-bar-bg { background: #EDF2F7; border-radius: 4px; height: 6px; }
.dim-score-bar-fill { height: 6px; border-radius: 4px; background: linear-gradient(90deg,#164A41,#E07B39); }
.dim-score-val { font-size: 12px; font-weight: 600; color: #164A41; width: 34px; text-align: right; margin-left: 10px; }

/* ── Compensation table ── */
.comp-table { width: 100%; border-collapse: collapse; }
.comp-table tr { border-bottom: 1px solid #EDF2F7; }
.comp-table tr:last-child { border-bottom: none; }
.comp-table td { padding: 8px 10px; font-size: 13px; }
.comp-table td:first-child { color: #64748B; font-weight: 500; }
.comp-table td:last-child { color: #164A41; font-weight: 600; text-align: right; }
.comp-table .comp-hi td { background: #EFF7F4; }
.comp-table .comp-hi td:last-child { color: #E07B39; }

/* ── Hint text under sliders ── */
.sub-dim-hint { font-size: 11.5px; color: #94A3B8; margin-bottom: 4px; font-style: italic; }

/* ── Comment area ── */
.comment-header {
    font-size: 11px; font-weight: 700; letter-spacing: 0.08em; text-transform: uppercase;
    color: #94A3B8; margin-bottom: 6px;
}

/* ── Computed score chip ── */
.computed-chip {
    background: #EFF7F4; border-radius: 8px; padding: 10px 14px;
    border: 1px solid #B8D8D2; font-size: 13.5px; margin: 6px 0 8px 0;
}

/* ── Domain heading + prompt ── */
.domain-heading {
    font-size: 30px !important;
    font-weight: 700 !important;
    color: #164A41 !important;
    margin: 32px 0 6px 0 !important;
    letter-spacing: -0.3px;
    line-height: 1.2;
}
.domain-prompt {
    font-size: 15.5px !important;
    color: #64748B !important;
    margin: 0 0 18px 0 !important;
    line-height: 1.55;
}

/* ── Anchor radio blocks ── */
div[data-testid="stRadio"] > label { font-size: 14px !important; font-weight: 600 !important; color: #164A41 !important; margin-bottom: 10px !important; }
div[data-testid="stRadio"] > div {
    display: flex !important;
    flex-direction: column !important;
    gap: 6px !important;
    width: 100% !important;
}
div[data-testid="stRadio"] > div > label {
    background: #F5FAF8 !important;
    border: 1px solid #D8EBE7 !important;
    border-radius: 8px !important;
    padding: 10px 14px !important;
    cursor: pointer !important;
    transition: background 0.15s, border-color 0.15s !important;
    width: 100% !important;
    box-sizing: border-box !important;
    align-items: flex-start !important;
}
div[data-testid="stRadio"] > div > label:hover {
    background: #EBF5F1 !important;
    border-color: #164A41 !important;
}
div[data-testid="stRadio"] > div > label:has(input:checked) {
    background: #E8F4F1 !important;
    border-color: #164A41 !important;
    border-left: 3px solid #164A41 !important;
}
div[data-testid="stRadio"] > div > label p { font-size: 14.5px !important; color: #334155 !important; line-height: 1.55 !important; margin: 0 !important; }

/* ── Streamlit widget overrides ── */
div[data-testid="stSelectSlider"] label { font-size: 15px !important; font-weight: 500 !important; color: #334155 !important; }
.stTextArea textarea { font-size: 14.5px !important; border-color: #D8EBE7 !important; border-radius: 8px !important; }
.stTextInput input { font-size: 14.5px !important; border-color: #D8EBE7 !important; border-radius: 8px !important; }
.stDateInput input { font-size: 14.5px !important; border-radius: 8px !important; }
.stCaption, [data-testid="stCaptionContainer"] p { font-size: 14px !important; }
.sub-dim-hint { font-size: 12.5px !important; }

/* Primary button (main area) */
.main .stButton > button[kind="primary"] {
    background: linear-gradient(135deg,#164A41 0%,#1D5C4E 100%) !important;
    border: none !important; border-radius: 8px !important;
    font-weight: 600 !important; font-size: 14px !important; color: #FFFFFF !important;
}
.stDownloadButton > button {
    background: linear-gradient(135deg,#E07B39 0%,#F5A462 100%) !important;
    border: none !important; border-radius: 8px !important;
    font-weight: 600 !important; color: #FFFFFF !important;
}
button[data-baseweb="tab"] { font-size: 13px !important; font-weight: 500 !important; }
details summary p { font-size: 13.5px !important; font-weight: 600 !important; color: #164A41 !important; }

/* Hide Streamlit chrome */
#MainMenu, footer, header { visibility: hidden; }
.stDeployButton { display: none; }
[data-testid="stDataFrame"] { border-radius: 10px !important; overflow: hidden; }
</style>
"""

# ──────────────────────────────────────────────────────────────────────────────
# SCORING LOGIC
# ──────────────────────────────────────────────────────────────────────────────
def calc_ic(subs):
    if subs.count(5) >= 2: return 5
    if subs.count(4) >= 2: return 4
    return round(sum(subs) / len(subs))

def calculate_scores():
    s = st.session_state
    tc = sum(s[k] for k in ["tc_legal","tc_data","tc_strategy","tc_leadership","tc_transformational"]) / 5
    ic_subs = [s[k] for k in ["bc_ic_cp","bc_ic_cs","bc_ic_team","bc_ic_org"]]
    ic = calc_ic(ic_subs)
    ic_n, fr_n = ic/5, s["bc_freq"]/5
    co_n, cf_n = s["bc_cons"]/5, s["bc_conf"]/5
    bc = round((ic_n**0.3 * fr_n**0.25 * co_n**0.25 * cf_n**0.2)*5, 2) if all(v>0 for v in [ic_n,fr_n,co_n,cf_n]) else 0.0
    mental = (sum(s[k] for k in ["ef_conc","ef_prob","ef_info","ef_multi","ef_switch"]) / 5) * 0.5
    emot   = (sum(s[k] for k in ["ef_own","ef_oth","ef_conf","ef_press"]) / 4) * 0.5
    effort = mental + emot
    pc   = sum(s[k] for k in ["pc_cred","pc_rel","pc_org"]) / 3
    wc   = sum(s[k] for k in ["wc_sched","wc_travel","wc_social"]) / 3
    resp = sum(s[k] for k in ["resp_scope","resp_auto","resp_rev","resp_dec"]) / 4
    raw  = tc*0.125 + bc*0.125 + effort*0.125 + pc*0.25 + wc*0.125 + resp*0.25
    return {"tc":tc,"bc":bc,"ic":ic,"effort":effort,"mental":mental,"emot":emot,
            "pc":pc,"wc":wc,"resp":resp,"raw":raw,"final":math.floor(raw*10)/10}

def lookup_level(score):
    return min(PAY_STRUCTURE.items(), key=lambda x: abs(x[1]["score"]-score))[0]

# ──────────────────────────────────────────────────────────────────────────────
# SHARED UI HELPERS
# ──────────────────────────────────────────────────────────────────────────────
def page_header(title, subtitle, badge=None):
    b = badge or title.upper()
    subtitle_html = f'<p class="page-subtitle">{subtitle}</p>' if subtitle else ""
    st.markdown(f"""
<div class="page-header">
  <div class="page-header-badge">{b}</div>
  <h1 class="page-title">{title}</h1>
  {subtitle_html}
</div>""", unsafe_allow_html=True)

def info_box(text):
    st.markdown(f'<div class="info-panel">{text}</div>', unsafe_allow_html=True)

def _reasoning_expander(score_key, subject):
    """Optional free-text reasoning, labelled with the sub-dimension subject."""
    cmt_key = f"comment_{score_key}"
    wkey = f"_wc_{cmt_key}"
    if wkey not in st.session_state:
        st.session_state[wkey] = st.session_state.get(cmt_key, "")
    has_text = bool(st.session_state.get(cmt_key, ""))
    label = f"Reasoning for {subject} (optional)"
    with st.expander(label, expanded=has_text):
        st.text_area(
            label,
            key=wkey,
            height=80,
            placeholder=f"Why did you select this score for {subject}? Add the reasoning here.",
            label_visibility="collapsed",
        )
    st.session_state[cmt_key] = st.session_state[wkey]

def score_slider(label, key, hint=None):
    """Slider + per-sub-dimension reasoning expander.

    Storage key (e.g. 'tc_legal') is persistent; widget key (_w_*) is what
    Streamlit uses for the rendered widget — kept separate so values survive
    page navigation.
    """
    lo = min_score(key)
    wkey = f"_w_{key}"
    if wkey not in st.session_state:
        st.session_state[wkey] = st.session_state.get(key, lo)
    if hint:
        st.markdown(f'<div class="sub-dim-hint">{hint}</div>', unsafe_allow_html=True)
    val = st.select_slider(label, options=list(range(lo, 6)),
                           format_func=lambda x: SCORE_LABELS[x], key=wkey)
    st.session_state[key] = val
    _reasoning_expander(key, label)
    return val

def anchor_radio(domain, key, prompt, anchors):
    """Domain heading + prompt + radio + per-sub-dimension reasoning expander.

    `anchors` is a list of (score, level, description) tuples covering 0–5.
    Sub-dimensions whose minimum is 1 simply skip the score-0 option.
    """
    lo = min_score(key)
    wkey = f"_w_{key}"
    if wkey not in st.session_state:
        st.session_state[wkey] = st.session_state.get(key, lo)
    st.markdown(f'<h2 class="domain-heading">{domain}</h2>', unsafe_allow_html=True)
    if prompt:
        st.markdown(f'<p class="domain-prompt">{prompt}</p>', unsafe_allow_html=True)
    val = st.radio(
        domain,
        options=list(range(lo, 6)),
        format_func=lambda x: f"{anchors[x][0]} - {anchors[x][1]}. {anchors[x][2]}",
        key=wkey,
        label_visibility="collapsed",
    )
    st.session_state[key] = val
    _reasoning_expander(key, domain)
    return val

def comment_box(key, placeholder="Add comments, justification or context…"):
    """Standalone text area (used for the overall reasoning on the Results page)."""
    wkey = f"_wc_{key}"
    if wkey not in st.session_state:
        st.session_state[wkey] = st.session_state.get(key, "")
    st.markdown('<div class="comment-header">Comments</div>', unsafe_allow_html=True)
    val = st.text_area("", key=wkey, height=90, placeholder=placeholder,
                       label_visibility="collapsed")
    st.session_state[key] = val

def score_bar_html(score, max_score=5):
    pct = max(0, min(100, score / max_score * 100))
    return f"""<div style="display:flex;align-items:center;gap:10px;">
  <div style="flex:1;background:#EDF2F7;border-radius:4px;height:6px;overflow:hidden;">
    <div style="width:{pct:.0f}%;height:6px;border-radius:4px;background:linear-gradient(90deg,#164A41,#E07B39);"></div>
  </div>
  <span style="font-size:13px;font-weight:600;color:#164A41;min-width:32px;text-align:right;">{score:.2f}</span>
</div>"""

def dimension_row(name, score, weight_label):
    pct = max(0, min(100, score / 5 * 100))
    st.markdown(f"""<div class="dim-row">
  <div class="dim-name">{name}</div>
  <div class="dim-weight">{weight_label}</div>
  <div class="dim-score-bar"><div class="dim-score-bar-bg">
    <div class="dim-score-bar-fill" style="width:{pct:.0f}%"></div>
  </div></div>
  <div class="dim-score-val">{score:.2f}</div>
</div>""", unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────────────────────────
# PAGES
# ──────────────────────────────────────────────────────────────────────────────
def _text(label, key, **kwargs):
    """Text input with persistent storage key (separate from widget key)."""
    wkey = f"_wi_{key}"
    if wkey not in st.session_state:
        st.session_state[wkey] = st.session_state.get(key, "")
    val = st.text_input(label, key=wkey, **kwargs)
    st.session_state[key] = val
    return val

def _textarea(label, key, **kwargs):
    wkey = f"_wt_{key}"
    if wkey not in st.session_state:
        st.session_state[wkey] = st.session_state.get(key, "")
    val = st.text_area(label, key=wkey, **kwargs)
    st.session_state[key] = val
    return val

def page_job_info():
    page_header("Job Information", "Basic details about the evaluation.", "New Evaluation")
    st.markdown('<div class="card"><div class="card-title">Evaluation Details</div>', unsafe_allow_html=True)
    _text("Employee / Candidate Name", "job_name")
    _text("Evaluator", "evaluator")
    st.date_input("Evaluation Date", key="eval_date")
    st.markdown('</div>', unsafe_allow_html=True)

    info_box("""<strong>How to use this tool</strong><br>
Score each sub-dimension from <strong>0 to 5</strong>. Scores are saved as you go — navigate freely between
pages and your inputs will be retained. The <strong>Results page</strong> shows the final weighted score,
the pay level, and the full compensation package, with an Excel export.""")

    dims = [
        ("Competency",           "Technical + Behavioural", "25%"),
        ("Professional Capital", "Trust and credibility",   "25%"),
        ("Responsibility",       "Accountability",          "25%"),
        ("Effort",               "Mental + Emotional",      "12.5%"),
        ("Working Conditions",   "Role context",            "12.5%"),
    ]
    cards = "".join(
        f"""<div style="background:#F5FAF8;border:1px solid #D8EBE7;border-radius:10px;padding:14px 16px;">
  <div style="font-size:11px;font-weight:700;letter-spacing:0.07em;text-transform:uppercase;color:#164A41;">{name}</div>
  <div style="font-size:11px;color:#94A3B8;margin-top:3px;">{sub}</div>
  <div style="font-size:18px;font-weight:700;color:#164A41;margin-top:8px;">{wt}</div>
</div>"""
        for name, sub, wt in dims
    )
    st.markdown(
        f'<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:12px;margin-top:4px;">{cards}</div>',
        unsafe_allow_html=True,
    )


TECH_ANCHORS = {
    "Legal": {
        "prompt": "How would you estimate the level of your legal expertise and the ability to apply legal reasoning to new and unfamiliar business and organisational questions?",
        "anchors": [
            (0, "Aptitude",                "I believe to have strong analytical and legal reasoning potential. I understand core legal logic conceptually, but have not yet applied it in a professional, real-life context."),
            (1, "Foundation",              "I have applied legal reasoning in a professional, real-life context. Although not yet an independent legal advisor, with the proper guidance I can deliver advice and take positions in standard contexts."),
            (2, "Independent advisor",     "I can independently analyse legal questions and translate law into business-relevant advice. I can deliver advice and take positions without any supervision or guidance."),
            (3, "Senior advisor",          "I can act as a trusted legal advisor for clients. I can deliver advice and take positions in complex, multidimensional contexts independently."),
            (4, "Authority",               "I can define a position in complex, high-impact situations and/or in completely new and unfamiliar contexts. Accordingly, I am able to shape legal frameworks and can act as a reference within the practice."),
            (5, "Thought leader",          "I cannot only be considered a reference within the practice internally, but also in the market. As such, I am able to develop new methodologies and influence industry thinking."),
        ],
    },
    "Data": {
        "prompt": "How would you estimate your data literacy and your ability to build structured analyses, models, and computations?",
        "anchors": [
            (0, "Aptitude",                "I believe to have strong quantitative and analytical aptitude. I understand basic data, mathematical, and logical concepts, but have not yet applied them in a professional, real-life context."),
            (1, "Foundation",              "I can perform basic calculations and use simple formulas. With the proper guidance, I can translate business logic into Excel (or another calculation framework) and follow predefined models."),
            (2, "Independent professional","I can build structured calculation models, perform multi-step computations, and check consistency and logic. I apply quantitative reasoning independently."),
            (3, "Advanced professional",   "I can design complex computational models, integrate multiple datasets, and handle assumptions, sensitivities, and scenarios. I combine computation with analytical judgment."),
            (4, "Expert",                  "I can design data and computation architectures and handle statistical, financial, or algorithmic complexity. I validate and challenge models both conceptually and technically."),
            (5, "Thought leader",          "I design and define advanced computational and analytical methodologies. I influence how data analysis is done across the firm and operate well beyond tool-level expertise."),
        ],
    },
    "Strategy": {
        "prompt": "How would you estimate your strategic thinking and your ability to integrate legal and quantitative insights into strategic recommendations?",
        "anchors": [
            (0, "Aptitude",                "I am curious about how organisations create value and I understand basic business logic. I can follow strategic reasoning, but have not yet applied it independently in a professional context."),
            (1, "Foundation",              "I understand basic strategy concepts and frameworks. With the proper guidance, I can apply standard strategy tools, structure information for analysis, and communicate strategic concepts clearly."),
            (2, "Independent professional","I can independently conduct structured strategic analyses, formulate coherent problem statements, and synthesise qualitative and quantitative insights to support decision-making."),
            (3, "Advanced specialist",     "I can design strategic analyses for complex contexts, challenge assumptions, evaluate options with clear trade-off logic, and integrate market, organisational, financial, and human factors."),
            (4, "Authority",               "I can define strategic direction for business units or organisations, design strategic planning and governance processes, and advise top decision-makers on long-term implications."),
            (5, "Thought leader",          "I develop new strategic concepts and methodologies. I influence organisational and industry strategic thinking and integrate strategy with transformation, culture, and governance."),
        ],
    },
    "Leadership": {
        "prompt": "How would you estimate your ability to enable Stratarius and its people to perform — through expertise, judgment, and example rather than hierarchical authority? This dimension applies to Stratarius itself, not to client engagements.",
        "anchors": [
            (0, "Aptitude",                "I am open to responsibility and growth within Stratarius, and I am aware of my own strengths and development areas. I accept guidance and supervision constructively, but have not yet demonstrated leadership of others inside the firm."),
            (1, "Self-leadership",         "I take ownership of my own work, quality, and deadlines at Stratarius. I act reliably and consistently, seek and apply feedback from colleagues, and demonstrate accountability for my own outputs and decisions."),
            (2, "Informal expert leader",  "I support Stratarius colleagues through content-based problem-solving and share my expertise proactively across the firm. I influence peers through professional reasoning and example, not authority."),
            (3, "Team leader",             "I can lead Stratarius teams, projects, or internal workstreams based on subject-matter authority. I guide colleagues' performance through coaching and resolve internal conflicts using professional judgment and credibility."),
            (4, "Organisational leader",   "I can lead multiple Stratarius teams or major internal units through professional authority. I shape the firm's culture and professional standards, and I build leadership and expert capability in fellow Stratarians."),
            (5, "Institutional leader",    "I shape Stratarius' identity, purpose, and long-term direction. I act as a moral, cultural, and professional reference point inside the firm, and I represent Stratarius externally as a trusted authority."),
        ],
    },
    "Transformational": {
        "prompt": "How would you estimate your capability to reshape Stratarius itself — its business model, organisation, and ways of working — over time? This dimension applies to Stratarius, not to transformation work done for external clients.",
        "anchors": [
            (0, "Aptitude",                "I understand that change and uncertainty are inherent to running a consultancy like Stratarius. I can function in evolving internal contexts with guidance, and I am open to learning and adjustment."),
            (1, "Adaptive professional",   "I can operate effectively when Stratarius itself is changing or ambiguous. I adjust my approach as the firm's strategy or organisation evolves and I maintain my performance during internal transitions."),
            (2, "Independent change operator", "I can anticipate the impact of strategic or organisational change on my own work and immediate Stratarius environment. I reframe internal problems as contexts evolve and help colleagues navigate uncertainty."),
            (3, "Transformation lead",     "I can design and lead internal change initiatives at Stratarius — translating firm strategy into transition paths, aligning colleagues during transformation, and managing systemic interdependencies inside the firm."),
            (4, "Strategic transformation leader", "I can shape large-scale transformations of Stratarius — redesigning the firm's operating model, service portfolio, or organisation. I balance stability and change and act as the internal reference authority on transformation."),
            (5, "System architect",        "I redefine how Stratarius itself adapts and evolves as a business. I reshape the firm's business model and create new operating paradigms, drawing on deep technical expertise and market knowledge."),
        ],
    },
}

def page_technical():
    page_header("Technical Competency",
                "Multidisciplinary expertise across five domains · Weight: 12.5% · Equally important as behavioural competency",
                "Technical")
    info_box("At Stratarius, <strong>leadership belongs to technical capability</strong> — because leadership without expertise has no legitimacy. Score each domain independently. The <strong>Legal</strong> domain has a minimum score of 1 since Stratarius is fundamentally a legal advisory firm; the other four domains can score 0. <strong>When in doubt, select the lower level.</strong>")

    keys = ["tc_legal", "tc_data", "tc_strategy", "tc_leadership", "tc_transformational"]
    for key, (domain, meta) in zip(keys, TECH_ANCHORS.items()):
        anchor_radio(domain, key, meta["prompt"], meta["anchors"])

    tc = sum(st.session_state[k] for k in keys) / 5
    st.markdown(f'<div style="margin:4px 0 16px 0;">{score_bar_html(tc)}</div>', unsafe_allow_html=True)
    st.caption(f"Technical Competency score: **{tc:.2f} / 5** (arithmetic average)")


BC_ANCHORS = {
    "bc_ic_cp": {
        "domain": "Complexity / ambiguity of the client problem",
        "prompt": "How would you describe the typical complexity and ambiguity of the client problems you engage with? (Does the role mainly answer questions, or define them?)",
        "anchors": [
            (0, "Not exposed",       "In my role I am not directly exposed to client problems; I work on internal tasks, analysis, or preparation only."),
            (1, "Clearly defined",   "The client problems I work on are clearly defined, scoped, and already translated into concrete questions."),
            (2, "Mostly defined",    "The client problems I work on are mostly defined; only minor clarification is needed before I can engage."),
            (3, "Partially defined", "The client problems I work on are partially defined; I need to do meaningful structuring and prioritisation myself."),
            (4, "Largely unstructured","The client problems I work on are largely unstructured; framing the question is a major part of the work."),
            (5, "Ill-defined",       "The client problems I work on are ill-defined or emerging; the core question must first be discovered."),
        ],
    },
    "bc_ic_cs": {
        "domain": "Complexity / ambiguity of the client system",
        "prompt": "How complex is the client organisation and stakeholder environment you typically operate in? (How hard is it to know who really decides?)",
        "anchors": [
            (0, "Not exposed",      "In my role I am not directly exposed to client stakeholder systems."),
            (1, "Simple structure", "I work with a single decision-maker in a simple structure, with limited internal dynamics."),
            (2, "Aligned",          "I work with a few stakeholders whose interests are mostly aligned."),
            (3, "Mixed priorities", "I work with several stakeholders who have different priorities."),
            (4, "Many stakeholders","I work with many stakeholders across functions or levels of the client organisation."),
            (5, "Politically complex","I operate in complex client organisations with fragmented ownership, politics, or unclear authority."),
        ],
    },
    "bc_ic_team": {
        "domain": "Team interaction",
        "prompt": "How much does your role require you to integrate and coordinate within the project team? (Are you mainly an individual contributor, or a team integrator?)",
        "anchors": [
            (0, "Solo",          "My role is fully solo; I have effectively no project-team interaction."),
            (1, "Independent",   "I work largely independently with only limited coordination."),
            (2, "Occasional",    "I coordinate occasionally with one or two team members."),
            (3, "Regular",       "I regularly align and integrate with several team members."),
            (4, "Active",        "I actively coordinate and synchronise the contributions of team members."),
            (5, "Main integrator","I act as the main integrator of the team's expertise."),
        ],
    },
    "bc_ic_org": {
        "domain": "Organisational interaction",
        "prompt": "How much does your role require you to operate across organisational boundaries and authority lines — business/technical leads, internal governance, firm interface? (Do you follow organisational structures, or connect them?)",
        "anchors": [
            (0, "Within team",     "I operate entirely within a single team and do not cross organisational boundaries."),
            (1, "Single line",     "I operate within one clear organisational line at Stratarius."),
            (2, "Occasional",      "I occasionally interact across one organisational boundary."),
            (3, "Regular",         "I regularly work across multiple internal functions or roles."),
            (4, "Frequent balancing","I frequently balance competing organisational interests."),
            (5, "Bridge",          "I act as a structural bridge between organisational logics (e.g. business vs. technical, commercial vs. professional)."),
        ],
    },
    "bc_freq": {
        "domain": "Frequency of interpersonal interaction",
        "prompt": "What share of your typical working time is spent in interpersonal interaction (internal or external), regardless of how complex that interaction is?",
        "anchors": [
            (0, "Negligible","My role involves essentially no interpersonal interaction."),
            (1, "Rare",      "Interpersonal interaction is incidental (<15% of my time). My work is mostly individual and task-focused."),
            (2, "Occasional","Interaction is required but not dominant (~15–30% of my time)."),
            (3, "Regular",   "Interaction is a normal part of my work (~30–50% of my time), with frequent alignment, collaboration, and feedback."),
            (4, "Predominant","Interaction dominates most of my working time (~50–70%) — discussions, workshops, advisory conversations."),
            (5, "Continuous","Interaction is the core of my role (>70% of my time). My role exists primarily to interact, influence, align, and communicate."),
        ],
    },
    "bc_cons": {
        "domain": "Consequence of interpersonal failures",
        "prompt": "If communication, alignment, or relationship management fails in your role, how serious are the consequences for Stratarius or its clients? (Not about how often failures occur — about risk exposure.)",
        "anchors": [
            (0, "No exposure","My role has no meaningful exposure to interpersonal failure risk."),
            (1, "Negligible", "Failures cause only minor inconvenience and are easily corrected (small rework, minor clarification)."),
            (2, "Limited",    "Failures cause short delays or local inefficiencies (task rework, limited stakeholder irritation)."),
            (3, "Noticeable", "Failures have a visible impact on project outcomes or stakeholder confidence (project delays, formal correction needed)."),
            (4, "Serious",    "Failures materially affect results, relationships, or credibility — management intervention may be required."),
            (5, "Critical",   "Failures have severe or long-lasting impact on business, reputation, or strategic relationships (loss of client, major reputational damage)."),
        ],
    },
    "bc_conf": {
        "domain": "Conflict and resistance management",
        "prompt": "How much does your role structurally require you to deal with disagreement, resistance, opposition, or emotionally charged situations to deliver results? (About the environment, not personal style.)",
        "anchors": [
            (0, "None",       "My role does not require dealing with disagreement or resistance."),
            (1, "Minimal",    "My interactions are mostly cooperative and aligned; resistance is rare."),
            (2, "Occasional", "Minor disagreements or professional differences of opinion occur but are easily resolved."),
            (3, "Regular",    "Constructive disagreement and resistance are a normal part of my work (challenging assumptions, negotiating trade-offs)."),
            (4, "Frequent",   "Strong resistance, emotional reactions, or political tension are common (change resistance, defensive stakeholders, difficult negotiations)."),
            (5, "Systemic",   "My role operates in an environment where resistance and conflict are persistent and structurally embedded (transformation, power shifts, high-stakes negotiations)."),
        ],
    },
}

EFFORT_ANCHORS = {
    "ef_conc": {
        "domain": "Concentration and focus required",
        "prompt": "How much sustained attention does your role typically require?",
        "anchors": [
            (0, "Minimal",        "My role requires only minimal sustained attention."),
            (1, "Short bursts",   "My role requires short periods of attention; interruptions are harmless."),
            (2, "Regular focus",  "My role requires regular focus, but frequent breaks are possible."),
            (3, "Sustained",      "My role requires sustained focus for meaningful periods of time."),
            (4, "Deep concentration","My role requires long periods of deep concentration."),
            (5, "Continuous",     "My role requires continuous high-intensity focus with little tolerance for error."),
        ],
    },
    "ef_prob": {
        "domain": "Complexity of problem-solving",
        "prompt": "How complex are the problems your role requires you to solve?",
        "anchors": [
            (0, "None",            "My role does not require independent problem-solving."),
            (1, "Routine",         "The problems I solve are routine and rule-based."),
            (2, "Standard judgment","The problems I solve require standard professional judgment."),
            (3, "Mixed",           "The problems I solve are of mixed complexity with some ambiguity."),
            (4, "High complexity", "The problems I solve are highly complex and multi-variable."),
            (5, "Novel / systemic","The problems I solve are novel, systemic, or ill-defined."),
        ],
    },
    "ef_info": {
        "domain": "Amount of information to process and retain",
        "prompt": "How much information does your role require you to process and retain?",
        "anchors": [
            (0, "Minimal",   "My role requires processing only minimal information."),
            (1, "Limited",   "My role involves limited information that is easy to retain."),
            (2, "Moderate",  "My role involves a moderate volume of information."),
            (3, "Significant","My role requires me to handle significant information that needs to be structured."),
            (4, "Large",     "My role requires me to handle large volumes of information across multiple domains."),
            (5, "Very large","My role requires me to handle very large, dense, and interdependent information sets."),
        ],
    },
    "ef_multi": {
        "domain": "Multitasking demands",
        "prompt": "How much does your role require you to manage multiple tasks in parallel?",
        "anchors": [
            (0, "None",          "My role is fully sequential; one task at a time."),
            (1, "Single-task",   "I focus on a single task most of the time."),
            (2, "Occasional",    "I handle occasional parallel tasks."),
            (3, "Regular",       "I regularly multitask."),
            (4, "Frequent",      "I frequently manage multiple tasks simultaneously."),
            (5, "Continuous",    "I multitask continuously across competing priorities."),
        ],
    },
    "ef_switch": {
        "domain": "Switching roles and contexts",
        "prompt": "How often does your role require you to switch between different contexts, roles, or cognitive modes?",
        "anchors": [
            (0, "None",       "My role has a single, consistent context."),
            (1, "Rare",       "I rarely switch between contexts."),
            (2, "Occasional", "I occasionally switch between contexts."),
            (3, "Regular",    "I regularly switch between different work contexts."),
            (4, "Frequent",   "I frequently switch between very different roles or cognitive modes."),
            (5, "Continuous", "I continuously switch across highly different cognitive modes throughout the day."),
        ],
    },
    "ef_own": {
        "domain": "Regulating my own emotions",
        "prompt": "How much does your role environment require you to regulate your own emotions (independent of your personal style)?",
        "anchors": [
            (0, "None",       "My role places no demands on emotional self-regulation."),
            (1, "Rare",       "My emotional neutrality is rarely challenged."),
            (2, "Occasional", "I occasionally need to remain calm or positive under pressure."),
            (3, "Regular",    "I regularly need to manage frustration, stress, or disappointment."),
            (4, "Frequent",   "I frequently need to suppress strong emotions to remain professional."),
            (5, "Continuous", "My role requires continuous emotional self-regulation under demanding conditions."),
        ],
    },
    "ef_oth": {
        "domain": "Managing others' emotions",
        "prompt": "How much does your role require you to manage the emotional states of clients, colleagues, or stakeholders?",
        "anchors": [
            (0, "None",        "My role does not involve dealing with others' emotional reactions."),
            (1, "Rare",        "I rarely deal with others' emotional reactions."),
            (2, "Occasional",  "I occasionally reassure or support others."),
            (3, "Regular",     "I regularly help others manage uncertainty or tension."),
            (4, "Frequent",    "I frequently manage strong emotional reactions from others."),
            (5, "Continuous",  "I constantly work with anxiety, resistance, or emotional dependency in others."),
        ],
    },
    "ef_conf": {
        "domain": "Dealing with conflict, complaints, and distress",
        "prompt": "How frequently and intensely does your role expose you to conflict, complaints, or distressed stakeholders?",
        "anchors": [
            (0, "None",       "My role does not involve conflict, complaints, or distress."),
            (1, "Rare",       "I am rarely exposed to complaints or distress."),
            (2, "Occasional", "I deal with occasional complaints or mild dissatisfaction."),
            (3, "Regular",    "I am regularly exposed to conflict or dissatisfaction."),
            (4, "Frequent",   "I am frequently exposed to serious complaints or distress."),
            (5, "Continuous", "I am continuously exposed to high emotional tension, conflict, or distress."),
        ],
    },
    "ef_press": {
        "domain": "Maintaining professional demeanour under pressure",
        "prompt": "How much sustained pressure does your role place on your ability to maintain a professional demeanour?",
        "anchors": [
            (0, "None",       "My role does not involve pressure situations."),
            (1, "Rare",       "Pressure situations are rare."),
            (2, "Limited",    "I experience limited pressure with easy recovery."),
            (3, "Regular",    "I face regular pressure that requires emotional control."),
            (4, "Frequent",   "I face high-pressure situations frequently."),
            (5, "Sustained",  "My role involves sustained high pressure with constant emotional restraint required."),
        ],
    },
}

PC_ANCHORS = {
    "pc_cred": {
        "domain": "Professional credibility",
        "prompt": "If you express a professional opinion, how much weight does it naturally carry? (Earned trust in professional judgment.)",
        "anchors": [
            (0, "None yet",     "I am a professional starter; my judgment is still being trained and requires proactive input, review, and validation by others."),
            (1, "Low",          "My credibility is mainly role-based, not person-based; my judgment is not yet trusted independently and is rarely accepted without confirmation."),
            (2, "Emerging",     "My credibility is developing but still context-dependent; my judgment is trusted in limited or routine situations, but I still need reinforcement by more senior professionals."),
            (3, "Established",  "My judgment is broadly trusted in relevant contexts; I am recognised as a reliable professional by peers and clients, and my advice is normally accepted without escalation."),
            (4, "Strong",       "I am widely trusted as a professional authority; my judgment influences decisions beyond my formal responsibility and I am frequently consulted."),
            (5, "Exceptional",  "I am a reference authority in my field; my judgment shapes thinking, standards, or direction and my credibility extends beyond Stratarius."),
        ],
    },
    "pc_rel": {
        "domain": "Relational capital (network)",
        "prompt": "To what extent do you have documented, recurring, and trusted professional relationships that can realistically be activated for Stratarius?",
        "anchors": [
            (0, "None",        "I have no owned or trusted professional relationships at all."),
            (1, "Minimal",     "I have no relationships that can be activated in the Stratarius context and no external professional visibility."),
            (2, "Limited",     "I have a small number of professional relationships with limited activation potential, and limited or early external visibility."),
            (3, "Established", "I have several owned professional relationships that can be activated in Stratarius, with clear evidence of repeat stakeholder trust and regular industry participation."),
            (4, "Strong",      "I have a broad and deep network of trusted professional relationships; stakeholders actively seek me out for advice or continuity."),
            (5, "Exceptional", "I have an extensive, high-trust professional network that materially influences access and opportunities, with stakeholder loyalty across employers."),
        ],
    },
    "pc_org": {
        "domain": "Organisational capital (continuity)",
        "prompt": "To what extent has durable trust, continuity, and relational embeddedness been built between you, Stratarius, and Stratarius' clients? (New joiners — also experienced ones — start at 0.)",
        "anchors": [
            (0, "None yet",    "I am a new joiner; internal relationships are still being established and no client continuity has been built through Stratarius."),
            (1, "Minimal",     "I have a limited internal network and trust, no meaningful client continuity through Stratarius, and limited familiarity with the firm's culture and informal networks."),
            (2, "Emerging",    "I have basic internal relationships and credibility, limited client continuity within Stratarius, and a growing understanding of collaboration patterns."),
            (3, "Established", "I have solid internal trust across multiple stakeholders, clear client continuity within Stratarius projects, and good cultural and organisational embeddedness."),
            (4, "Strong",      "I have deep internal trust and strong relational positioning; clients associate Stratarius continuity with me and I am a stabilising reference point."),
            (5, "Exceptional", "I am an institutional anchor for Stratarius — long-term client relationships are materially tied to my presence and I bridge trust across generations of clients and colleagues."),
        ],
    },
}

WC_ANCHORS = {
    "wc_sched": {
        "domain": "Schedule demands",
        "prompt": "How predictable and controllable is your typical working schedule?",
        "anchors": [
            (0, "Fully predictable",  "My schedule is fully predictable and stable; no unexpected demands."),
            (1, "Highly predictable", "My schedule is highly predictable with very limited variation."),
            (2, "Mostly predictable", "My schedule is mostly predictable, with limited peaks."),
            (3, "Mixed",              "My schedule has mixed predictability — both stable and unpredictable periods."),
            (4, "Frequently unpredictable","My schedule is frequently unpredictable or involves extended hours."),
            (5, "Highly disruptive",  "My schedule is highly unpredictable, irregular, or disruptive."),
        ],
    },
    "wc_travel": {
        "domain": "Travel demands",
        "prompt": "How much travel does your role typically require? (Frequency, distance, and burden.)",
        "anchors": [
            (0, "None",       "My role involves no travel at all."),
            (1, "Negligible", "My role involves no or negligible travel."),
            (2, "Occasional", "My role involves occasional local travel."),
            (3, "Regular",    "My role involves regular regional travel."),
            (4, "Frequent",   "My role involves frequent national or international travel."),
            (5, "Continuous", "My role involves continuous or highly demanding travel."),
        ],
    },
    "wc_social": {
        "domain": "Social and organisational environment",
        "prompt": "What is the social and organisational climate you work in — pressure, politics, and the level of team support versus self-reliance?",
        "anchors": [
            (0, "Highly supportive","My environment is highly supportive with abundant team support and low pressure."),
            (1, "Stable",          "My environment is stable, supportive, and low-pressure."),
            (2, "Generally positive","My environment is generally positive."),
            (3, "Mixed",           "My environment is mixed, with occasional tension."),
            (4, "Pressured",       "My environment is frequently high-pressure or politically sensitive."),
            (5, "Structurally pressured","My environment is structurally high-pressure and conflict-prone."),
        ],
    },
}

RESP_ANCHORS = {
    "resp_scope": {
        "domain": "Scope of impact",
        "prompt": "If your role consistently made poor decisions, how far would the damage realistically spread? (Breadth and significance of outcomes affected — not authority.)",
        "anchors": [
            (0, "None",            "My role has no decision impact beyond my own day-to-day execution."),
            (1, "Local",           "My impact is limited to individual tasks or small deliverables."),
            (2, "Team-level",      "My impact affects a small team or project component."),
            (3, "Project / function","My impact affects entire projects, major workstreams, or a defined function."),
            (4, "Organisational / client","My impact materially affects Stratarius performance or client outcomes."),
            (5, "Strategic",       "My impact affects long-term strategy, positioning, or sustainability."),
        ],
    },
    "resp_auto": {
        "domain": "Autonomy and decision-making authority",
        "prompt": "To what extent does your role independently make decisions and commit Stratarius, clients, or projects without requiring approval? (About who decides, not the complexity of the decision.)",
        "anchors": [
            (0, "None",     "I have no decision-making role; I execute instructions only."),
            (1, "Prescribed","My decisions are fully prescribed or approved by others; I have no independent decision power."),
            (2, "Limited",  "I make minor operational decisions within clear boundaries; they are local, reversible, and often reviewed."),
            (3, "Moderate", "I independently make professional decisions within a defined scope and am trusted to decide in my own domain without routine approval."),
            (4, "High",     "I make significant decisions affecting projects, clients, or teams; approvals are the exception, not the norm."),
            (5, "Full",     "I have authority to commit Stratarius or client direction; I shape outcomes through independent decisions."),
        ],
    },
    "resp_rev": {
        "domain": "Reversibility and risk",
        "prompt": "If your role makes a wrong decision, how easy is it to undo — and how bad is the damage if it cannot be undone?",
        "anchors": [
            (0, "No risk",         "My decisions carry no meaningful risk and are trivial to reverse."),
            (1, "Fully reversible","My decisions are easily corrected with little or no negative impact."),
            (2, "Mostly reversible","My decisions may cause inconvenience or delay but are still correctable."),
            (3, "Partially reversible","Some of my decisions have lasting consequences, though mitigation is possible."),
            (4, "Largely irreversible","Many of my decisions have serious and lasting consequences; errors may damage client trust, financial outcomes, or reputation."),
            (5, "Irreversible",    "My decisions have long-term or permanent strategic, financial, or reputational impact; mistakes can materially harm Stratarius' future positioning."),
        ],
    },
    "resp_dec": {
        "domain": "Decision complexity and frequency",
        "prompt": "How frequent and complex are the decisions your role requires you to take? (Viewed from the organisation's perspective, not personal mental effort.)",
        "anchors": [
            (0, "None",             "My role does not involve material decision-making."),
            (1, "Rare and simple",  "Decisions are infrequent and straightforward, following clear rules."),
            (2, "Occasional",       "Decisions are occasional and of low complexity, with limited variables."),
            (3, "Regular and moderate","Decisions are a normal part of my role and involve multiple considerations."),
            (4, "Frequent and complex","Decisions are frequent and involve high ambiguity or multiple stakeholders."),
            (5, "Continuous and complex","Decision-making is constant and highly complex — ambiguous, multi-variable, high-impact decisions."),
        ],
    },
}


def _render_anchor_block(key, dct):
    meta = dct[key]
    anchor_radio(meta["domain"], key, meta["prompt"], meta["anchors"])


def page_behavioural():
    page_header("Behavioural Competency",
                "Interaction complexity, frequency, consequence & conflict · Weight: 12.5% · Equally important as technical competency",
                "Behavioural")
    info_box("The overall BC score uses a <strong>weighted geometric mean</strong> — a very low score on any single dimension significantly reduces the total. <strong>When in doubt, select the lower level.</strong>")
    for key in ["bc_ic_cp", "bc_ic_cs", "bc_ic_team", "bc_ic_org"]:
        _render_anchor_block(key, BC_ANCHORS)
    ic = calc_ic([st.session_state[k] for k in ["bc_ic_cp","bc_ic_cs","bc_ic_team","bc_ic_org"]])
    st.markdown(f'<div class="computed-chip"><span style="color:#94A3B8;">Interaction Complexity (computed):</span> <strong style="color:#E07B39;font-size:17px;">{ic}</strong> / 5  <span style="color:#94A3B8;font-size:12px;">— ceiling rule: 2 or more 5s → 5; 2 or more 4s → 4; otherwise rounded average</span></div>', unsafe_allow_html=True)
    for key in ["bc_freq", "bc_cons", "bc_conf"]:
        _render_anchor_block(key, BC_ANCHORS)
    ic_n = ic/5; fr_n = st.session_state["bc_freq"]/5
    co_n = st.session_state["bc_cons"]/5; cf_n = st.session_state["bc_conf"]/5
    bc = round((ic_n**0.3*fr_n**0.25*co_n**0.25*cf_n**0.2)*5,2) if all(v>0 for v in [ic_n,fr_n,co_n,cf_n]) else 0.0
    st.markdown(f'<div style="margin:18px 0 16px 0;">{score_bar_html(bc)}</div>', unsafe_allow_html=True)
    st.caption(f"Behavioural Competency score: **{bc:.2f} / 5** (weighted geometric mean × 5)")


def page_effort():
    page_header("Effort",
                "Mental burden (50%) + Emotional burden (50%) · Weight: 12.5%",
                "Effort")
    info_box("Effort recognises the <strong>cognitive and emotional demands</strong> of the role. Mental and emotional effort contribute equally (50% each). <strong>When in doubt, select the lower level.</strong>")
    for key in ["ef_conc", "ef_prob", "ef_info", "ef_multi", "ef_switch"]:
        _render_anchor_block(key, EFFORT_ANCHORS)
    mental = (sum(st.session_state[k] for k in ["ef_conc","ef_prob","ef_info","ef_multi","ef_switch"]) / 5) * 0.5
    st.markdown(f'<div style="margin:14px 0 4px 0;">{score_bar_html(mental, max_score=2.5)}</div>', unsafe_allow_html=True)
    st.caption(f"Mental effort contribution: **{mental:.2f}** (avg × 0.5)")
    for key in ["ef_own", "ef_oth", "ef_conf", "ef_press"]:
        _render_anchor_block(key, EFFORT_ANCHORS)
    emot = (sum(st.session_state[k] for k in ["ef_own","ef_oth","ef_conf","ef_press"]) / 4) * 0.5
    st.markdown(f'<div style="margin:14px 0 4px 0;">{score_bar_html(emot, max_score=2.5)}</div>', unsafe_allow_html=True)
    st.caption(f"Emotional effort contribution: **{emot:.2f}** (avg × 0.5)")
    effort = mental + emot
    st.markdown(f'<div style="margin:14px 0 16px 0;">{score_bar_html(effort)}</div>', unsafe_allow_html=True)
    st.caption(f"Total Effort score: **{effort:.2f} / 5**")


def page_professional():
    page_header("Professional Capital",
                "Trust, credibility and accumulated capital · Weight: 25% · One of the two highest-weighted dimensions",
                "Professional Capital")
    info_box("Professional Capital is one of the two <strong>highest-weighted dimensions (25%)</strong> because Stratarius is a customer-intimate consultancy where trust and credibility are absolutely core. <strong>When in doubt, select the lower level.</strong>")
    for key in ["pc_cred", "pc_rel", "pc_org"]:
        _render_anchor_block(key, PC_ANCHORS)
    pc = sum(st.session_state[k] for k in ["pc_cred","pc_rel","pc_org"]) / 3
    st.markdown(f'<div style="margin:18px 0 16px 0;">{score_bar_html(pc)}</div>', unsafe_allow_html=True)
    st.caption(f"Professional Capital score: **{pc:.2f} / 5** (arithmetic average)")


def page_working():
    page_header("Working Conditions",
                "Schedule, travel and social environment · Weight: 12.5%",
                "Working Conditions")
    info_box("Working conditions recognise the <strong>context</strong> in which the role operates. Higher scores indicate more demanding or less favourable conditions. <strong>When in doubt, select the lower level.</strong>")
    for key in ["wc_sched", "wc_travel", "wc_social"]:
        _render_anchor_block(key, WC_ANCHORS)
    wc = sum(st.session_state[k] for k in ["wc_sched","wc_travel","wc_social"]) / 3
    st.markdown(f'<div style="margin:18px 0 16px 0;">{score_bar_html(wc)}</div>', unsafe_allow_html=True)
    st.caption(f"Working Conditions score: **{wc:.2f} / 5** (arithmetic average)")


def page_responsibility():
    page_header("Level of Responsibility",
                "Scope, autonomy, risk and decision complexity · Weight: 25% · One of the two highest-weighted dimensions",
                "Responsibility")
    info_box("Responsibility is the other highest-weighted dimension (25%) — Stratarius values <strong>accountability</strong> and the ability to own outcomes independently. <strong>When in doubt, select the lower level.</strong>")
    for key in ["resp_scope", "resp_auto", "resp_rev", "resp_dec"]:
        _render_anchor_block(key, RESP_ANCHORS)
    resp = sum(st.session_state[k] for k in ["resp_scope","resp_auto","resp_rev","resp_dec"]) / 4
    st.markdown(f'<div style="margin:18px 0 16px 0;">{score_bar_html(resp)}</div>', unsafe_allow_html=True)
    st.caption(f"Responsibility score: **{resp:.2f} / 5** (arithmetic average)")


def page_results():
    page_header("Results & Pay Proposal",
                "Final score, pay level and full compensation package", "Results")
    sc = calculate_scores()
    cat = lookup_level(sc["final"])
    pay = PAY_STRUCTURE[cat]

    # Top KPI row
    kpis = [("Final Score", f"{sc['final']:.1f}", True),
            ("Raw Score",   f"{sc['raw']:.3f}", False),
            ("Pay Level",   cat, False),
            ("Base Salary", f"€{pay['salary']:,.0f}/mo", False)]
    cols = st.columns(4)
    for col, (label, val, hi) in zip(cols, kpis):
        extra = "metric-highlight" if hi else ""
        col.markdown(f"""<div class="metric-card {extra}">
  <div class="metric-label">{label}</div>
  <div class="metric-value">{val}</div>
</div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    col_left, col_right = st.columns([3, 2])

    with col_left:
        st.markdown('<div class="card"><div class="card-title">Score Breakdown</div>', unsafe_allow_html=True)
        dims = [
            ("Technical Competency",    sc["tc"],     "12.5%"),
            ("Behavioural Competency",  sc["bc"],     "12.5%"),
            ("Effort",                  sc["effort"], "12.5%"),
            ("Professional Capital",    sc["pc"],     "25%"),
            ("Working Conditions",      sc["wc"],     "12.5%"),
            ("Level of Responsibility", sc["resp"],   "25%"),
        ]
        for name, score, wt in dims:
            dimension_row(name, score, wt)
        st.markdown('</div>', unsafe_allow_html=True)

    with col_right:
        annual = pay["salary"] * 13.92
        st.markdown(f"""<div class="card">
  <div class="card-title">Pay Level</div>
  <div style="text-align:center;margin-bottom:16px;">
    <div style="font-size:10px;font-weight:600;letter-spacing:0.1em;text-transform:uppercase;color:#94A3B8;margin-bottom:6px;">Category</div>
    <div class="pay-level-badge">{cat}</div>
    <div style="font-size:11px;color:#94A3B8;margin-top:6px;">Score threshold: {pay['score']}</div>
  </div>
  <table class="comp-table">
    <tr class="comp-hi"><td>Base monthly (gross)</td><td>€ {pay['salary']:,.2f}</td></tr>
    <tr><td>Annual (× 13.92)</td><td>€ {annual:,.0f}</td></tr>
    <tr class="comp-hi"><td>Mobility budget</td><td>€ {pay['mobility']:,} /mo</td></tr>
    <tr><td>Home work allowance</td><td>€ 150 /mo</td></tr>
    <tr><td>Meal vouchers</td><td>€ 10 / day</td></tr>
    <tr><td>Ecovouchers</td><td>€ 250 /yr</td></tr>
    <tr><td>Yearly premium</td><td>€ 330.84</td></tr>
    <tr><td>Collective bonus</td><td>up to €3,700 net/yr</td></tr>
    <tr><td>Paid days off</td><td>45 days /yr</td></tr>
  </table>
</div>""", unsafe_allow_html=True)

    with st.expander("Detailed weighted calculation"):
        weights = [0.125, 0.125, 0.125, 0.25, 0.125, 0.25]
        scores_vals = [sc["tc"],sc["bc"],sc["effort"],sc["pc"],sc["wc"],sc["resp"]]
        dim_names = ["Technical Competency","Behavioural Competency","Effort",
                     "Professional Capital","Working Conditions","Level of Responsibility"]
        rows = [{"Dimension":n,"Score":round(s,3),"Weight":w,"Weighted":round(s*w,4)}
                for n,s,w in zip(dim_names,scores_vals,weights)]
        rows.append({"Dimension":"TOTAL","Score":"","Weight":sum(weights),"Weighted":round(sc["raw"],4)})
        st.dataframe(pd.DataFrame(rows), hide_index=True, use_container_width=True)

    st.markdown('<hr style="border:none;border-top:1px solid #EDF2F7;margin:20px 0;">', unsafe_allow_html=True)
    comment_box("overall_comments", "Overall observations and rationale for the proposed pay level…")
    st.markdown('<hr style="border:none;border-top:1px solid #EDF2F7;margin:20px 0;">', unsafe_allow_html=True)

    st.markdown('<div class="card-title" style="padding-bottom:10px;border-bottom:1px solid #EDF2F7;margin-bottom:14px;">Export</div>', unsafe_allow_html=True)
    if st.button("Generate Excel Report", type="primary", use_container_width=True):
        buf = export_excel(sc)
        candidate = st.session_state.get("job_name","evaluation").replace(" ","_")
        st.download_button("Download Excel Report", data=buf,
                           file_name=f"stratarius_job_evaluation_{candidate}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)


def page_info():
    page_header("Reward Strategy", "", "Reference")
    tab1, tab2 = st.tabs(["Philosophy", "Scoring Logic"])

    with tab1:
        st.markdown("""
> *"At Stratarius we primarily value (and therefore reward) technical and behavioural competence,
> professional credibility and responsibility, while recognizing effort and working conditions as
> important but secondary job factors."*

| Category | Message |
|--|--|
| **Competency** | We value capability — technical and behavioural equally |
| **Professional Capital** | We value trust and credibility |
| **Responsibility** | We value accountability |
| **Effort** | We recognize burden |
| **Working Conditions** | We recognize context |

### Governance

The pay level is based on the individual's capabilities *(looking back)* and the job demands *(looking forward)*. The process is **fully transparent** and the subject of an open discussion. How the dimensions and subdimensions are scored to determine the value of the work is subject to dialogue. The outcome of the applied principles and the corresponding pay level, however, are non-negotiable.
""")

    with tab2:
        st.markdown("""
### Dimension Weights

| Dimension | Weight | Why |
|--|--|--|
| Technical Competency | 12.5% | Core capability |
| Behavioural Competency | 12.5% | Core capability — equal to technical |
| **Professional Capital** | **25%** | Trust and credibility — most critical |
| **Level of Responsibility** | **25%** | Accountability — most critical |
| Effort | 12.5% | Recognized but secondary |
| Working Conditions | 12.5% | Context — recognized but secondary |

### Scoring Methods

- **Technical Competency:** Arithmetic average of 5 domains. Scores 0–5, except **Legal** which has a minimum of 1 (Stratarius is fundamentally a legal advisory firm). Only TC and PC allow scores of 0.
- **Behavioural Competency:** Weighted geometric mean — `(IC^0.3 × Freq^0.25 × Cons^0.25 × Conf^0.2) × 5`. Scores 1–5.
- **Interaction Complexity rule:** 2+ scores of 5 → IC = 5 | 2+ scores of 4 → IC = 4 | else: rounded average.
- **Effort:** `AVERAGE(mental subs) × 0.5 + AVERAGE(emotional subs) × 0.5`. Scores 1–5.
- **Professional Capital:** Arithmetic average of 3 dimensions. Scores 0–5.
- **Working Conditions & Responsibility:** Arithmetic average. Scores 1–5.
- **Final score:** Weighted sum → rounded *down* to 1 decimal.
""")


# ──────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ──────────────────────────────────────────────────────────────────────────────
def export_excel(sc):
    s = st.session_state
    wb = Workbook()
    navy="164A41"; orange="E07B39"; light="EFF7F4"; mid="1D5C4E"
    thin = Side(style="thin", color="D1DDED")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(wrap_text=True, vertical="top")

    def hdr(ws, r, c, v, fg=navy, size=11):
        cell = ws.cell(r, c, v)
        cell.fill = PatternFill("solid", fgColor=fg)
        cell.font = Font(color="FFFFFF", bold=True, size=size)
        cell.alignment = ctr; cell.border = bdr; return cell

    def dat(ws, r, c, v, bold=False, bg=None, align=None):
        cell = ws.cell(r, c, v)
        if bold:  cell.font  = Font(bold=True, size=10)
        if bg:    cell.fill  = PatternFill("solid", fgColor=bg)
        cell.alignment = align or wrap; cell.border = bdr; return cell

    # Sheet 1: Job Info
    ws1 = wb.active; ws1.title = "Job Info"
    ws1.column_dimensions["A"].width = 28; ws1.column_dimensions["B"].width = 44
    ws1.row_dimensions[1].height = 30
    hdr(ws1,1,1,"JOB ARCHITECTURE EVALUATION — STRATARIUS",navy,13); ws1.merge_cells("A1:B1")
    for i,(k,v) in enumerate([
        ("Employee / Candidate", s.get("job_name","")),
        ("Evaluator", s.get("evaluator","")),
        ("Evaluation Date", str(s.get("eval_date",""))),
    ], start=2):
        dat(ws1,i,1,k,bold=True,bg=light); dat(ws1,i,2,v)

    # Sheet 2: Scoring Details
    ws2 = wb.create_sheet("Scoring Details")
    ws2.column_dimensions["A"].width=42; ws2.column_dimensions["B"].width=10
    ws2.column_dimensions["C"].width=12; ws2.column_dimensions["D"].width=12; ws2.column_dimensions["E"].width=50
    hdr(ws2,1,1,"Sub-dimension"); hdr(ws2,1,2,"Score"); hdr(ws2,1,3,"Weight"); hdr(ws2,1,4,"Dim Score"); hdr(ws2,1,5,"Comments")
    row = [2]

    def add_dim(name, score, wt):
        r = row[0]
        hdr(ws2,r,1,name,mid); dat(ws2,r,2,"")
        dat(ws2,r,3,wt,align=ctr); dat(ws2,r,4,round(score,3),bold=True,align=ctr); dat(ws2,r,5,"",align=wrap)
        row[0] += 1

    def add_sub(name, score, score_key, wt=""):
        r = row[0]
        comment = s.get(f"comment_{score_key}","") if score_key else ""
        dat(ws2,r,1,f"    {name}",bg=light); dat(ws2,r,2,score,align=ctr)
        dat(ws2,r,3,wt,align=ctr); dat(ws2,r,4,""); dat(ws2,r,5,comment,align=wrap)
        row[0] += 1

    add_dim("1. Technical Competency", sc["tc"], "12.5%")
    for k,lbl,w in [("tc_legal","Legal","20%"),("tc_data","Data","20%"),("tc_strategy","Strategy","20%"),("tc_leadership","Leadership","20%"),("tc_transformational","Transformational","20%")]:
        add_sub(lbl, s.get(k,0), k, w)

    add_dim("2. Behavioural Competency", sc["bc"], "12.5%")
    for k,lbl in [("bc_ic_cp","IC - Client problem"),("bc_ic_cs","IC - Client system"),("bc_ic_team","IC - Team interaction"),("bc_ic_org","IC - Org interaction")]:
        add_sub(lbl, s.get(k,0), k)
    add_sub(f"Interaction Complexity (computed: {sc['ic']})", sc['ic'], None, "30%")
    for k,lbl,w in [("bc_freq","Frequency","25%"),("bc_cons","Consequence","25%"),("bc_conf","Conflict","20%")]:
        add_sub(lbl, s.get(k,0), k, w)

    add_dim("3. Effort", sc["effort"], "12.5%")
    for k,lbl in [("ef_conc","Mental: Concentration"),("ef_prob","Mental: Problem-solving"),("ef_info","Mental: Information"),("ef_multi","Mental: Multitasking"),("ef_switch","Mental: Role switching")]:
        add_sub(lbl, s.get(k,0), k)
    for k,lbl in [("ef_own","Emotional: Own emotions"),("ef_oth","Emotional: Others emotions"),("ef_conf","Emotional: Conflict"),("ef_press","Emotional: Pressure")]:
        add_sub(lbl, s.get(k,0), k)

    add_dim("4. Professional Capital", sc["pc"], "25%")
    for k,lbl in [("pc_cred","Professional credibility"),("pc_rel","Relational capital"),("pc_org","Organizational capital")]:
        add_sub(lbl, s.get(k,0), k, "1/3")

    add_dim("5. Working Conditions", sc["wc"], "12.5%")
    for k,lbl in [("wc_sched","Schedule demands"),("wc_travel","Travel demands"),("wc_social","Social environment")]:
        add_sub(lbl, s.get(k,0), k, "33%")

    add_dim("6. Level of Responsibility", sc["resp"], "25%")
    for k,lbl in [("resp_scope","Scope of impact"),("resp_auto","Autonomy and decision authority"),("resp_rev","Reversibility and risk"),("resp_dec","Decision complexity")]:
        add_sub(lbl, s.get(k,0), k, "25%")

    # Sheet 3: Results
    ws3 = wb.create_sheet("Results & Pay Proposal")
    ws3.column_dimensions["A"].width=36; ws3.column_dimensions["B"].width=24; ws3.column_dimensions["C"].width=18
    ws3.row_dimensions[1].height=30
    hdr(ws3,1,1,"RESULTS & PAY PROPOSAL — STRATARIUS",navy,13); ws3.merge_cells("A1:C1")
    r=2; hdr(ws3,r,1,"Dimension",mid); hdr(ws3,r,2,"Score",mid); hdr(ws3,r,3,"Weight",mid); r+=1
    for name,score,wt in [("Technical Competency",sc["tc"],"12.5%"),("Behavioural Competency",sc["bc"],"12.5%"),
                           ("Effort",sc["effort"],"12.5%"),("Professional Capital",sc["pc"],"25%"),
                           ("Working Conditions",sc["wc"],"12.5%"),("Level of Responsibility",sc["resp"],"25%")]:
        dat(ws3,r,1,name); dat(ws3,r,2,round(score,3),align=ctr); dat(ws3,r,3,wt,align=ctr); r+=1
    dat(ws3,r,1,"Raw Final Score",bold=True,bg=light); dat(ws3,r,2,round(sc["raw"],4),bold=True,align=ctr); dat(ws3,r,3,"",bg=light); r+=1
    dat(ws3,r,1,"FINAL SCORE (rounded down)",bold=True)
    c2=ws3.cell(r,2,sc["final"]); c2.font=Font(bold=True,size=14,color=orange); c2.fill=PatternFill("solid",fgColor="FFF3E8"); c2.border=bdr; c2.alignment=ctr
    dat(ws3,r,3,""); r+=2
    cat = lookup_level(sc["final"]); pay = PAY_STRUCTURE[cat]
    hdr(ws3,r,1,"PAY PROPOSAL",mid); ws3.merge_cells(f"A{r}:C{r}"); r+=1
    for label,value in [
        ("Pay Level", cat), ("Base Monthly Salary (gross)", f"€ {pay['salary']:,.2f}"),
        ("Annual Salary (x 13.92)", f"€ {pay['salary']*13.92:,.2f}"),
        ("Mobility Budget", f"€ {pay['mobility']:,} / month"), ("Home Work Allowance", "€ 150 / month"),
        ("Meal Vouchers", "€ 10 / worked day"), ("Ecovouchers", "€ 250 / year"),
        ("Yearly Premium (2026)", "€ 330.84"), ("Collective Bonus", "Up to €3,700 net / year (FTE)"),
        ("Paid Days Off", "45 days / year (FTE)"), ("Overall Comments", s.get("overall_comments","")),
    ]:
        dat(ws3,r,1,label,bold=True,bg=light); c2=ws3.cell(r,2,value); c2.border=bdr; c2.alignment=wrap; ws3.merge_cells(f"B{r}:C{r}"); r+=1

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ──────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ──────────────────────────────────────────────────────────────────────────────
def _logo_data_url():
    """Read assets/stratarius-logo.png at runtime and embed as a data URL."""
    p = Path(__file__).with_name("assets") / "stratarius-logo.png"
    if not p.exists():
        return None
    return f"data:image/png;base64,{base64.b64encode(p.read_bytes()).decode()}"

def _inject_scroll_behavior(page_id: str):
    """First visit → scroll to top. Revisit → restore previous scroll position.

    Uses browser sessionStorage (persists within the same tab across Streamlit
    reruns) and window.parent to reach the actual Streamlit scroll container
    from inside the component iframe.
    """
    components.html(f"""
<script>
(function() {{
    var pageId   = {repr(page_id)};
    var posKey   = 'scroll_pos_'     + pageId;
    var visitKey = 'scroll_visited_' + pageId;

    function getScrollEl() {{
        try {{
            return window.parent.document.querySelector('[data-testid="stMain"]') ||
                   window.parent.document.querySelector('.main') ||
                   window.parent.document.documentElement;
        }} catch(e) {{ return null; }}
    }}

    function applyScroll() {{
        var el = getScrollEl();
        if (!el) return;
        if (!sessionStorage.getItem(visitKey)) {{
            el.scrollTop = 0;
            try {{ window.parent.scrollTo(0, 0); }} catch(e) {{}}
            sessionStorage.setItem(visitKey, '1');
        }} else {{
            var pos = parseInt(sessionStorage.getItem(posKey) || '0', 10);
            el.scrollTop = pos;
            try {{ window.parent.scrollTo(0, pos); }} catch(e) {{}}
        }}
    }}

    function saveScroll() {{
        try {{
            var el = getScrollEl();
            var pos = (el && el.scrollTop) ? el.scrollTop
                      : (window.parent.pageYOffset || 0);
            sessionStorage.setItem(posKey, pos);
        }} catch(e) {{}}
    }}

    // Apply after Streamlit content finishes painting
    setTimeout(applyScroll, 120);
    setTimeout(applyScroll, 400);

    // Persist scroll position while the user scrolls
    try {{ getScrollEl().addEventListener('scroll', saveScroll, {{passive: true}}); }} catch(e) {{}}
    try {{ window.parent.addEventListener('scroll', saveScroll, {{passive: true}}); }} catch(e) {{}}
}})();
</script>
""", height=0)


def render_sidebar():
    sc = calculate_scores()
    cat = lookup_level(sc["final"])
    pay = PAY_STRUCTURE[cat]
    pct = int(min(sc["final"] / 4.5 * 100, 100))
    current = st.session_state.get("_page", "job_info")

    with st.sidebar:
        logo_src = _logo_data_url()
        logo_img = f'<img src="{logo_src}" alt="Stratarius">' if logo_src else '<div style="font-size:22px;font-weight:700;color:#164A41;">Stratarius</div>'
        st.markdown(f"""<div class="sidebar-logo">
  {logo_img}
  <div class="sidebar-brand">Job Architecture Tool</div>
</div>""", unsafe_allow_html=True)

        for label, page_id in PAGES:
            is_active = (current == page_id)
            if st.button(label, key=f"nav_{page_id}",
                         use_container_width=True,
                         type="primary" if is_active else "secondary"):
                st.session_state["_page"] = page_id
                st.rerun()

        st.markdown(f"""<div class="score-preview">
  <div class="score-preview-label">Live Score</div>
  <div class="score-big">{sc['final']:.1f}</div>
  <div class="score-level">{cat}</div>
  <div class="score-salary">€{pay['salary']:,.0f} / month gross</div>
  <div class="score-bar-bg"><div class="score-bar-fill" style="width:{pct}%"></div></div>
</div>""", unsafe_allow_html=True)


# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────
def main():
    st.set_page_config(
        page_title="Job Architecture Scoring — Stratarius",
        page_icon="S",
        layout="wide",
        initial_sidebar_state="expanded",
    )
    st.markdown(STYLES, unsafe_allow_html=True)
    init_state()  # ← all defaults set once; widgets use key= only, no value= override

    render_sidebar()

    dispatch = {
        "job_info":       page_job_info,
        "technical":      page_technical,
        "behavioural":    page_behavioural,
        "effort":         page_effort,
        "professional":   page_professional,
        "working":        page_working,
        "responsibility": page_responsibility,
        "results":        page_results,
        "info":           page_info,
    }
    current_page = st.session_state["_page"]
    dispatch.get(current_page, page_job_info)()
    _inject_scroll_behavior(current_page)


if __name__ == "__main__":
    main()
